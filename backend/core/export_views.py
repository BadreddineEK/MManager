"""
Vues d'export — étape 11
=========================
Tous les exports nécessitent IsAuthenticated.
Les filtres (mosque, année, mois) sont passés en query params.

Endpoints :
  GET /api/export/families/excel/
  GET /api/export/families/pdf/
  GET /api/export/children/excel/
  GET /api/export/school-payments/excel/
  GET /api/export/school-payments/pdf/
  GET /api/export/members/excel/
  GET /api/export/members/pdf/
  GET /api/export/membership-payments/excel/
  GET /api/export/treasury/excel/
  GET /api/export/treasury/pdf/
"""
import io
from datetime import date

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework.permissions import IsAuthenticated

from core.permissions import HasMosquePermission
from rest_framework.views import APIView

from core.utils import get_mosque
from membership.models import Member, MembershipPayment
from school.models import Child, Family, SchoolPayment
from treasury.models import TreasuryTransaction

# ── Helpers styles Excel ──────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="3B3B8C")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL    = PatternFill("solid", fgColor="EDEDF5")


def _style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _excel_response(wb: Workbook, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _pdf_response(buf: io.BytesIO, filename: str) -> HttpResponse:
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _pdf_table_style():
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B3B8C")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EDEDF5")]),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])


def _pdf_header(story, title: str, mosque_name: str):
    styles = getSampleStyleSheet()
    story.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    story.append(Paragraph(f"{mosque_name} — généré le {date.today().strftime('%d/%m/%Y')}", styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))


# ── FAMILLES ─────────────────────────────────────────────────────────────────

class FamiliesExcelView(APIView):
    permission_classes = [IsAuthenticated, HasMosquePermission]

    def get(self, request):
        mosque = get_mosque(request)
        qs = Family.objects.filter(mosque=mosque).prefetch_related("children").order_by("primary_contact_name")

        wb = Workbook()
        ws = wb.active
        ws.title = "Familles"
        headers = ["ID", "Contact", "Téléphone 1", "Téléphone 2", "Email", "Nb enfants"]
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))

        for i, f in enumerate(qs, start=2):
            ws.append([f.id, f.primary_contact_name, f.phone1, f.phone2 or "", f.email or "", f.children.count()])
            if i % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = ALT_FILL

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

        return _excel_response(wb, f"familles_{date.today()}.xlsx")


class FamiliesPDFView(APIView):
    permission_classes = [IsAuthenticated, HasMosquePermission]

    def get(self, request):
        mosque = get_mosque(request)
        qs = Family.objects.filter(mosque=mosque).prefetch_related("children").order_by("primary_contact_name")

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm,
                                 topMargin=1.5*cm, bottomMargin=1*cm)
        story = []
        _pdf_header(story, "📋 Liste des familles", mosque.name if mosque else "")

        data = [["ID", "Contact", "Téléphone 1", "Téléphone 2", "Email", "Nb enfants"]]
        for f in qs:
            data.append([str(f.id), f.primary_contact_name, f.phone1, f.phone2 or "—", f.email or "—", str(f.children.count())])

        t = Table(data, colWidths=[1.5*cm, 5*cm, 3.5*cm, 3.5*cm, 5*cm, 2.5*cm])
        t.setStyle(_pdf_table_style())
        story.append(t)
        doc.build(story)
        return _pdf_response(buf, f"familles_{date.today()}.pdf")


# ── ENFANTS ──────────────────────────────────────────────────────────────────

class ChildrenExcelView(APIView):
    permission_classes = [IsAuthenticated, HasMosquePermission]

    def get(self, request):
        mosque = get_mosque(request)
        qs = Child.objects.filter(mosque=mosque).select_related("family").order_by("first_name")

        wb = Workbook()
        ws = wb.active
        ws.title = "Enfants"
        headers = ["ID", "Prénom", "Niveau", "Famille", "Date de naissance"]
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))

        for i, c in enumerate(qs, start=2):
            ws.append([
                c.id, c.first_name, c.level,
                c.family.primary_contact_name if c.family else "",
                c.birth_date.strftime("%d/%m/%Y") if c.birth_date else "",
            ])
            if i % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = ALT_FILL

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        return _excel_response(wb, f"enfants_{date.today()}.xlsx")


# ── PAIEMENTS ÉCOLE ──────────────────────────────────────────────────────────

class SchoolPaymentsExcelView(APIView):
    permission_classes = [IsAuthenticated, HasMosquePermission]

    def get(self, request):
        mosque = get_mosque(request)
        qs = (SchoolPayment.objects
              .filter(mosque=mosque)
              .select_related("child", "child__family")
              .order_by("-date"))

        wb = Workbook()
        ws = wb.active
        ws.title = "Paiements école"
        headers = ["ID", "Date", "Enfant", "Famille", "Montant (€)", "Méthode", "Notes"]
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))

        for i, p in enumerate(qs, start=2):
            ws.append([
                p.id,
                p.date.strftime("%d/%m/%Y"),
                f"{p.child.first_name}" if p.child else "—",
                p.child.family.primary_contact_name if p.child and p.child.family else "—",
                float(p.amount),
                p.method,
                p.note or "",
            ])
            if i % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = ALT_FILL

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

        return _excel_response(wb, f"paiements_ecole_{date.today()}.xlsx")


class SchoolPaymentsPDFView(APIView):
    permission_classes = [IsAuthenticated, HasMosquePermission]

    def get(self, request):
        mosque = get_mosque(request)
        qs = (SchoolPayment.objects
              .filter(mosque=mosque)
              .select_related("child", "child__family")
              .order_by("-date"))

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm,
                                 topMargin=1.5*cm, bottomMargin=1*cm)
        story = []
        total = sum(float(p.amount) for p in qs)
        _pdf_header(story, f"💰 Paiements école — Total : {total:.2f} €", mosque.name if mosque else "")

        data = [["Date", "Enfant", "Famille", "Montant (€)", "Méthode"]]
        for p in qs:
            data.append([
                p.date.strftime("%d/%m/%Y"),
                f"{p.child.first_name}" if p.child else "—",
                p.child.family.primary_contact_name if p.child and p.child.family else "—",
                f"{float(p.amount):.2f}",
                p.method,
            ])

        t = Table(data, colWidths=[2.5*cm, 5*cm, 5*cm, 3*cm, 3*cm])
        t.setStyle(_pdf_table_style())
        story.append(t)
        doc.build(story)
        return _pdf_response(buf, f"paiements_ecole_{date.today()}.pdf")


# ── ADHÉRENTS ────────────────────────────────────────────────────────────────

class MembersExcelView(APIView):
    permission_classes = [IsAuthenticated, HasMosquePermission]

    def get(self, request):
        mosque = get_mosque(request)
        qs = Member.objects.filter(mosque=mosque).order_by("full_name")

        wb = Workbook()
        ws = wb.active
        ws.title = "Adhérents"
        headers = ["ID", "Nom complet", "Email", "Téléphone", "Adresse"]
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))

        for i, m in enumerate(qs, start=2):
            ws.append([m.id, m.full_name, m.email or "", m.phone or "", m.address or ""])
            if i % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = ALT_FILL

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

        return _excel_response(wb, f"adherents_{date.today()}.xlsx")


class MembersPDFView(APIView):
    permission_classes = [IsAuthenticated, HasMosquePermission]

    def get(self, request):
        mosque = get_mosque(request)
        qs = Member.objects.filter(mosque=mosque).order_by("full_name")

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm,
                                 topMargin=1.5*cm, bottomMargin=1*cm)
        story = []
        _pdf_header(story, f"🤝 Liste des adhérents ({qs.count()})", mosque.name if mosque else "")

        data = [["Nom complet", "Email", "Téléphone", "Adresse"]]
        for m in qs:
            data.append([m.full_name, m.email or "—", m.phone or "—", m.address or "—"])

        t = Table(data, colWidths=[5*cm, 5*cm, 3.5*cm, 8*cm])
        t.setStyle(_pdf_table_style())
        story.append(t)
        doc.build(story)
        return _pdf_response(buf, f"adherents_{date.today()}.pdf")


# ── COTISATIONS ADHÉSION ──────────────────────────────────────────────────────

class MembershipPaymentsExcelView(APIView):
    permission_classes = [IsAuthenticated, HasMosquePermission]

    def get(self, request):
        mosque = get_mosque(request)
        qs = (MembershipPayment.objects
              .filter(mosque=mosque)
              .select_related("member", "membership_year")
              .order_by("-date"))

        wb = Workbook()
        ws = wb.active
        ws.title = "Cotisations"
        headers = ["ID", "Date", "Adhérent", "Année", "Montant (€)", "Méthode"]
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))

        for i, p in enumerate(qs, start=2):
            ws.append([
                p.id,
                p.date.strftime("%d/%m/%Y"),
                p.member.full_name if p.member else "—",
                p.membership_year.label if p.membership_year else "—",
                float(p.amount),
                p.method,
            ])
            if i % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = ALT_FILL

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

        return _excel_response(wb, f"cotisations_{date.today()}.xlsx")


# ── TRÉSORERIE ────────────────────────────────────────────────────────────────

class TreasuryExcelView(APIView):
    permission_classes = [IsAuthenticated, HasMosquePermission]

    def get(self, request):
        mosque = get_mosque(request)
        qs = TreasuryTransaction.objects.filter(mosque=mosque).order_by("-date")

        wb = Workbook()
        ws = wb.active
        ws.title = "Trésorerie"
        headers = ["ID", "Date", "Direction", "Catégorie", "Libellé", "Montant (€)", "Méthode"]
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))

        for i, t in enumerate(qs, start=2):
            ws.append([
                t.id,
                t.date.strftime("%d/%m/%Y"),
                "Entrée" if t.direction == "IN" else "Sortie",
                t.category,
                t.label,
                float(t.amount),
                t.method,
            ])
            fill = PatternFill("solid", fgColor="D1FAE5") if t.direction == "IN" else PatternFill("solid", fgColor="FEE2E2")
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = fill

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

        return _excel_response(wb, f"tresorerie_{date.today()}.xlsx")


class TreasuryPDFView(APIView):
    permission_classes = [IsAuthenticated, HasMosquePermission]

    def get(self, request):
        mosque = get_mosque(request)
        qs = TreasuryTransaction.objects.filter(mosque=mosque).order_by("-date")

        total_in  = sum(float(t.amount) for t in qs if t.direction == "IN")
        total_out = sum(float(t.amount) for t in qs if t.direction == "OUT")
        balance   = total_in - total_out

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm,
                                 topMargin=1.5*cm, bottomMargin=1*cm)
        story = []
        _pdf_header(
            story,
            f"🏦 Trésorerie — Entrées : {total_in:.2f} € | Sorties : {total_out:.2f} € | Solde : {balance:.2f} €",
            mosque.name if mosque else "",
        )

        data = [["Date", "Direction", "Catégorie", "Libellé", "Montant (€)", "Méthode"]]
        for t in qs:
            data.append([
                t.date.strftime("%d/%m/%Y"),
                "Entrée" if t.direction == "IN" else "Sortie",
                t.category,
                t.label,
                f"{float(t.amount):.2f}",
                t.method,
            ])

        style = _pdf_table_style()
        style.add("TEXTCOLOR", (1, 1), (1, -1), colors.HexColor("#059669"))  # direction col vert
        tbl = Table(data, colWidths=[2.5*cm, 2.5*cm, 3.5*cm, 7*cm, 3*cm, 3*cm])
        tbl.setStyle(style)
        story.append(tbl)
        doc.build(story)
        return _pdf_response(buf, f"tresorerie_{date.today()}.pdf")
